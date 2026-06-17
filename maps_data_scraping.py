import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd 

# Opening the Browser
driver=webdriver.Chrome()   
driver.maximize_window()

# Go on the google maps 
driver.get("https://www.google.com/maps")   
wait=WebDriverWait(driver,10)   


try:
# Add keyword in the search box
    search_box=wait.until(
        EC.presence_of_element_located(
            (By.XPATH, '//*[@id="ucc-1"]')
            
        )
    ) 

    search_box.send_keys("Software Companies in Noida") 
    search_box.send_keys(Keys.ENTER)    
    time.sleep(5)
    panel=driver.find_element(By.XPATH, ".//div[@role='feed']")
    print("Scrolling is starting please wait")  


    last_height=driver.execute_script("return arguments[0].scrollHeight",panel) 

    for _ in range(4):
        driver.execute_script("arguments[0].scrollTo(0,arguments[0].scrollHeight);",panel)   
        time.sleep(3)   
        new_height=driver.execute_script("return arguments[0].scrollHeight",panel)
        if new_height==last_height:
            break
        last_height=new_height
        
    wait.until(EC.presence_of_all_elements_located((By.XPATH, '//a[contains(@href, "/maps/place/")]')))
    companies=driver.find_elements(By.XPATH,'//a[contains(@href, "/maps/place/")]') 


    companies_list=[]
    seen_names=set()    
    print(f"\nTotal {len(companies)} links found")


    #     try:
    #         name=company.get_attribute("aria-label")    
    #         if name and name not in seen_names:
    #             companies_list.append({"Company Name":name})
    #             seen_names.add(name)    
    #     except Exception:
    #         continue
    
    for index, company in enumerate(companies,1):
        try:
            name=company.get_attribute("aria-label")
            if name and name not in seen_names:
                seen_names.add(name)
                
                driver.execute_script("arguments[0].click();",company)
                time.sleep(4)   
                
                phone="Not Available"   
                
                
            
                try:
                    # Right panel ke saare text blocks uthaye
                    detail_divs = driver.find_elements(By.XPATH, '//div[contains(@class, "fontBodyMedium")] | //span[contains(@class, "fontBodyMedium")]')
                    
                    for div in detail_divs:
                        text_val = div.text.strip() 
                        
                        # Phone number ki pehchan:
                        # 1. Usme digits hone chahiye.
                        # 2. Uske andar comma (,) nahi hona chahiye kyunki address me comma hota hai.
                        # 3. Us text ki total length 16 characters se choti honi chahiye (e.g., "+91 120 456 7890").
                        if text_val and any(char.isdigit() for char in text_val):
                            if len(text_val) <= 16 and "," not in text_val and "Noida" not in text_val:
                                if "open" not in text_val.lower() and "closes" not in text_val.lower() and "." not in text_val:
                                    phone = text_val
                                    break
                except Exception:
                    pass
                
                # print(f"{index}.Name:{name} | Phone: {phone}")  
                
                companies_list.append({
                    "Company Name": name,
                    "Phone Number": phone
                })
        except Exception:
            continue        


        if companies_list:
            df=pd.DataFrame(companies_list) 
            
    print(df)
    
    df.to_excel("Google_maps_data.xlsx", index=False)
    print("\n Data successfully saved in Google_maps_data.xlsx")

except Exception as e:
    print(f"Any error is occured:{e}")

    
finally:
    driver.quit()


# Designing the excel file

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Existing Excel file open karo
file_name = "Google_maps_data.xlsx"

wb = load_workbook(file_name)
ws = wb.active

# Style Definitions
header_font = Font(
    name="Calibri",
    size=12,
    bold=True,
    color="FFFFFF"
)

header_fill = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

data_font = Font(
    name="Calibri",
    size=11
)

thin_side = Side(
    style="thin",
    color="BFBFBF"
)

data_border = Border(
    left=thin_side,
    right=thin_side,
    top=thin_side,
    bottom=thin_side
)

center_align = Alignment(
    horizontal="center",
    vertical="center"
)

left_align = Alignment(
    horizontal="left",
    vertical="center"
)

# Header Design
ws.row_dimensions[1].height = 26

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = data_border

# Data Rows Design
for row in ws.iter_rows(
    min_row=2,
    max_row=ws.max_row,
    min_col=1,
    max_col=ws.max_column
):
    for cell in row:

        cell.font = data_font
        cell.border = data_border

        if cell.column == 2:
            cell.alignment = center_align
        else:
            cell.alignment = left_align

# Auto Width
for column in ws.columns:

    max_len = 0
    column_letter = column[0].column_letter

    for cell in column:
        try:
            if cell.value:
                max_len = max(
                    max_len,
                    len(str(cell.value))
                )
        except:
            pass

    ws.column_dimensions[column_letter].width = max(
        max_len + 4,
        15
    )

# Save
wb.save(file_name)

print("🎉 Excel formatting completed successfully!")